import csv
import os
from pathlib import Path
from openpyxl import load_workbook


class DataCaptureInterface(object):

    results_file = None
    experiment_type_mapper = {
        "nc": "No Cache",
        "nc_mg": "No Cache",
        "sc_dm": "Standard Cache - DM",
        "sc_dm_mg": "Standard Cache - DM",
        "sc_nway": "Standard Cache - Nway",
        "sc_nway_mg": "Standard Cache - Nway",
        "cc_dm": "Complex Cache - DM",
        "cc_dm_mg": "Complex Cache - DM",
        "cc_nway": "Complex Cache - Nway",
        "cc_nway_mg": "Complex Cache - Nway"
    }

    def __init__(self, results_file):
        self.results_file = Path(results_file)

    def result_present(self, benchmark_name, experiment_type):
        workbook = load_workbook(self.results_file)
        return any([x.value == benchmark_name for x in workbook.get_sheet_by_name(
            self.experiment_type_mapper[experiment_type]
        )["B"]])

    def mg_result_present(self, benchmark_name, experiment_type, results_folder):
        return any([x == "{}_{}_results.csv".format(benchmark_name, experiment_type) for x in os.listdir(results_folder)])

    def store_result(self, benchmark_name, experiment_type, results):
        workbook = load_workbook(self.results_file)
        sheet = workbook.get_sheet_by_name(self.experiment_type_mapper[experiment_type])
        filled = False
        # Initialise the counter to the first row under the headers
        counter = 4
        while not filled:
            if sheet["B{}".format(counter)].value is None:
                column = 2
                for element in [benchmark_name] + results:
                    sheet.cell(column=column, row=counter, value=element)
                    column += 1
                filled = True
            counter += 1
        workbook.save(self.results_file)

    def store_mg_result(self, benchmark_name, experiment_type, results, results_folder):
        headings = ["Memory Gap", "Counter when Gap Reported", "Addr Post Gap", "Counter with GNT Seen", "Start Time"]
        with open(Path(results_folder, "{}_{}_results.csv".format(benchmark_name, experiment_type)), "w") as fp:
            writer = csv.writer(fp)
            writer.writerow(headings)
            for row in results:
                writer.writerow(row)
